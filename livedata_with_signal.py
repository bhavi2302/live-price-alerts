import pandas as pd
import openpyxl
import os
import pyotp, time, pytz
from datetime import datetime, timezone
import threading
from SmartApi import SmartConnect
from SmartApi.smartWebSocketV2 import SmartWebSocketV2
from logzero import logger

# File paths
EXCEL_FILE_PATH = "C:/Users/bhave/Documents/Astrotrade108/stocks.xlsx"

# Print the current working directory
print("Current working directory:", os.getcwd())

# Check if the file exists
if not os.path.exists(EXCEL_FILE_PATH):
    print(f"{EXCEL_FILE_PATH} does not exist.")
    exit()

# Read the list of scripts from the Excel file
stocks_df = pd.read_excel(EXCEL_FILE_PATH)
if 'token' not in stocks_df.columns:
    print("The Excel file must contain a 'token' column.")
    exit()


apikey = 'kFa5C8jL'
username = 'B38590'
pwd = '2302'
token = 'EEHSUJJLWVFSZBRPEXVXKSFCLE'

# Create object of call
obj = SmartConnect(api_key=apikey)

# Login API call
print("Logging in...")
data = obj.generateSession(username, pwd, pyotp.TOTP(token).now())
refreshToken = data['data']['refreshToken']
AUTH_TOKEN = data['data']['jwtToken']
print("Logged in successfully!")

# Fetch the feedtoken
print("Fetching feed token...")
FEED_TOKEN = obj.getfeedToken()
print("Feed token fetched successfully!")

# Fetch User Profile
print("Fetching user profile...")
res = obj.getProfile(refreshToken) 
print(res['data']['exchanges'])

# Prepare the list of tokens
token_list = [
    {
        "exchangeType": 1,  # Assuming all tokens are from exchange type 1
        "tokens": list(stocks_df['token'].astype(str).values)
    }
]

# Define correlation_id and mode
correlation_id = "stocks_update"
mode = 1  # 1 for subscribe

sws = SmartWebSocketV2(AUTH_TOKEN, apikey, username, FEED_TOKEN, max_retry_attempt=5)

# Initialize an empty list to store the data
live_data = []

# Define your start and end dates here
start_date = datetime(2025, 1, 3, 10, 47, 0, tzinfo=pytz.timezone('Asia/Kolkata'))
end_date = datetime(2025, 1, 7, 15, 30, 0, tzinfo=pytz.timezone('Asia/Kolkata'))

# Function to append data to the Excel file every few seconds
def append_to_excel():
    if live_data:
        try:
            print("Appending to Excel...")  # Debugging statement
            df = pd.DataFrame(live_data)
            print("DataFrame created:\n", df)  # Debugging statement
            
            # Load the workbook and select the sheet
            wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
            sheet = wb.active
            
            # Create a dictionary to map tokens to their respective row index in the Excel sheet
            token_row_map = {str(cell.value): cell.row for cell in sheet['A'][1:]}  # Ensure token is a string
            print("Token row map:", token_row_map)  # Debugging statement
            
            # Fetch key levels for each token
            actual_high = sheet['E2'].value
            actual_low = sheet['F2'].value
            target_high = sheet['G2'].value
            target_low = sheet['H2'].value
            midpoints = [sheet[f'I{i}'].value for i in range(2, sheet.max_row+1)]
                
            # Write data to the appropriate rows
            for index, row in df.iterrows():
                token = str(row['token'])  # Ensure token is a string
                print(f"Processing token {token}")  # Debugging statement
                if token in token_row_map:
                    row_index = token_row_map[token]
                    ltp = row['last_traded_price']
                    print(f"Writing data for token {token} at row {row_index}")  # Debugging statement
                    sheet.cell(row=row_index, column=3, value=ltp)  # Column C for LTP
                    
                    # Calculate and append buy/sell signals based on all midpoints
                    buy_signal = "Buy" if abs(ltp - actual_low) <= 2 or abs(ltp - target_low) <= 2 or any(abs(ltp - mid) <= 2 for mid in midpoints) else ""
                    sell_signal = "Sell" if abs(ltp - actual_high) <= 2 or abs(ltp - target_high) <= 2 or any(abs(ltp - mid) <= 2 for mid in midpoints) else ""
                    
                    sheet.cell(row=row_index, column=10, value=buy_signal)  # Column J for Buy Signal
                    sheet.cell(row=row_index, column=11, value=sell_signal)  # Column K for Sell Signal
                else:
                    print(f"Token {token} not found in Excel sheet")  # Debugging statement
            
            # Save the workbook and close it properly
            wb.save(EXCEL_FILE_PATH)
            wb.close()
                
            logger.info("Data appended to stocks.xlsx")
            print("Data appended to stocks.xlsx")  # Debugging statement
            live_data.clear()
        except Exception as e:
            logger.error(f"Exception: {e}")
            print(f"Exception: {e}")

# Function to schedule the Excel update
def schedule_excel_update():
    print("Scheduling Excel update...")  # Debugging statement
    threading.Timer(60, schedule_excel_update).start()  # Adjust interval as needed
    append_to_excel()

# Start the schedule
print("Starting schedule...")
schedule_excel_update()

def on_data(wsapp, message):
    print("Received message: ", message)  # Debugging statement
    # Convert timestamp from milliseconds to seconds
    timestamp = message['exchange_timestamp'] / 1000  # Convert to seconds
    utc_time = datetime.fromtimestamp(timestamp, timezone.utc)  # Use timezone-aware object

    # Define the timezone for UTC +05:30
    timezone_india = pytz.timezone('Asia/Kolkata')  # 'Asia/Kolkata' is the timezone for UTC +05:30

    # Convert UTC time to UTC +05:30
    local_time = utc_time.astimezone(timezone_india)
    formatted_timestamp = local_time.strftime('%Y-%m-%d %H:%M:%S')
    print(f"Local time: {local_time}, Formatted timestamp: {formatted_timestamp}")  # Debugging statement

    # Append the data to the live_data list if within the specified date range
    if start_date <= local_time <= end_date:
        live_data.append({
            'exchange_type': message['exchange_type'],
            'token': message['token'],
            'last_traded_price': message['last_traded_price'] / 100  # Assuming this division by 100 is required for your specific case
        })
        print("Appended data to live_data list:", live_data[-1])  # Debugging statement
    else:
        print(f"Data received outside date range: {local_time}")  # Debugging statement

def on_open(wsapp):
    logger.info("WebSocket connection opened")
    print("WebSocket connection opened")  # Debugging statement
    sws.subscribe(correlation_id, mode, token_list)

def on_error(wsapp, error):
    logger.info(f"WebSocket error: {error}")
    print(f"WebSocket error: {error}")  # Debugging statement

def on_close(wsapp):
    logger.info("WebSocket connection closed")
    print("WebSocket connection closed")  # Debugging statement

def close_connection():
    sws.close_connection()

# Assign the callbacks.
sws.on_open = on_open
sws.on_data = on_data
sws.on_error = on_error
sws.on_close = on_close

print("Connecting to WebSocket...")
threading.Thread(target=sws.connect).start()
